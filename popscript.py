import openpyxl
import requests
import re
import os
import xmltodict
import datetime
from openpyxl.styles import Alignment

while True:
    try:
        name = input("Имя файла: ").replace(".xlsx","")
        # Проверка существования исходной таблицы
        if not os.path.exists(f"./{name}.xlsx"):
            print("Ошибка! Таблица не существует.")
            continue
        # Открытие исходной таблицы
        wb = openpyxl.load_workbook(filename=f'./{name}.xlsx')
        sheet = wb[wb.sheetnames[0]]

        # Удаление измененной ранее таблицы(если она есть)
        if os.path.exists(f"./{name}Edited.xlsx"):
            os.remove(f"./{name}Edited.xlsx")
        print("Таблица успешно загружена!")
        break
    except PermissionError:
        print("Ошибка! Возможно итоговая таблица уже где-то открыта.")
    except Exception as e:
        print(f"Ошибка! {e}")

# Создание новой таблицы
new_wb = openpyxl.Workbook()
new_wb.create_sheet('Лист1')
new_wb.remove(new_wb['Sheet'])
new_sheet = new_wb['Лист1']
for i,col in enumerate(["Города", "Даты", "Население", "Название страницы Википедии", "Ссылка на страницу Википедии"]):
    new_sheet.cell(column=i+1,row=1).value = col
new_wb.save(f"./{name}Edited.xlsx")

# Определение доп. данных
max_row = len(sheet["A"])
start_cell = "A3"
end_cell = f"A26"
col = 1
row = 2

# Открытие пустой таблицы
new_wb = openpyxl.load_workbook(filename=f'./{name}Edited.xlsx')
new_sheet = new_wb[new_wb.sheetnames[0]]

# Регулярное выражение для удаления ненужных символов в строке с населением
re_population = re.compile(r'(\d[\d ]*\d|\d+)')

for cells in sheet[f"{start_cell}:{end_cell}"]:
    if not cells[0].value: 
        # Пропуск пустых строк
        continue
    city_name = cells[0].value
    date = sheet.cell(column=3, row=cells[0].row).value
    wikipedia_data = requests.get(f"https://ru.wikipedia.org/w/api.php?format=xml&action=query&list=search&srwhat=text&srsearch={city_name.split(' ',1)[1]}")
    wiki_doc = xmltodict.parse(wikipedia_data.text)
    try:
        # Парсинг возможных страниц с населением
        search_pages = [page['@title'] for page in wiki_doc['api']['query']['search']['p'] if page['@title'].startswith(city_name.split(' ',1)[1]) and "(штат)" not in page['@title'] and 'район' not in page['@title']]
    except Exception:
        continue
    for page_title in search_pages:
        try:
            page_data = requests.get(f"https://ru.wikipedia.org/wiki/{page_title}")
            population_data = re_population.findall(page_data.text)
            population = int(population_data[0].replace(' ',''))
            wikipedia_link = f"https://ru.wikipedia.org/wiki/{page_title}"
            # Запись данных в новую таблицу
            new_sheet.cell(column=col, row=row).value = city_name
            new_sheet.cell(column=col+1, row=row).value = date
            new_sheet.cell(column=col+1, row=row).alignment = Alignment(horizontal='right')
            new_sheet.cell(column=col+2, row=row).value = population
            new_sheet.cell(column=col+2, row=row).number_format = "# ### ##0"
            new_sheet.cell(column=col+3, row=row).value = page_title
            new_sheet.cell(column=col+4, row=row).value = wikipedia_link
            row += 1
            break
        except Exception as e:
            continue
    else:
        # Если ни одна страница не подошла
        new_sheet.cell(column=col, row=row).value = city_name
        new_sheet.cell(column=col+1, row=row).value = date
        new_sheet.cell(column=col+1, row=row).alignment = Alignment(horizontal='right')
        new_sheet.cell(column=col+2, row=row).value = "Данные не найдены"
        row += 1

# Настройка выравнивания ячеек
new_sheet['A1'].alignment = Alignment(horizontal='center')
new_sheet['B1'].alignment = Alignment(horizontal='center')
new_sheet['C1'].alignment = Alignment(horizontal='center')
new_sheet['D1'].alignment = Alignment(horizontal='center')
new_sheet['E1'].alignment = Alignment(horizontal='center')

# Сохранение полученной таблицы
new_wb.save(f"./{name}Edited.xlsx")
print("Готово!")